"""
Output writer: takes the original input XLSX bytes and appends owner-name
columns (Owner 1, Owner 2, ...) to the right of the existing data. Also writes
a JSONL audit trail with the full SOS payload for debugging.
"""

import io
import json
import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from models import EntityRecord

logger = logging.getLogger("fdd-agent")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(size=10)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_xlsx_with_owners(original_bytes: bytes, records: list[EntityRecord]) -> bytes:
    """
    Open the input XLSX, append "Owner 1..N" columns after the last existing
    column, and write each record's owner names into the correct row (matched
    by EntityRecord.original_row_index).
    """
    wb = openpyxl.load_workbook(io.BytesIO(original_bytes))
    ws = wb.active

    max_owners = max((len(rec.people) for rec in records), default=0)
    if max_owners == 0:
        max_owners = 1  # at least one column so the operator sees the change

    first_new_col = ws.max_column + 1
    for i in range(max_owners):
        cell = ws.cell(row=1, column=first_new_col + i, value=f"Owner {i + 1}")
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(first_new_col + i)].width = 28

    by_row = {rec.original_row_index: rec for rec in records if rec.original_row_index > 0}

    for row_idx, rec in by_row.items():
        for i in range(max_owners):
            value = rec.people[i].name if i < len(rec.people) else ""
            cell = ws.cell(row=row_idx, column=first_new_col + i, value=value)
            cell.font = BODY_FONT
            cell.alignment = LEFT

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def write_audit_trail(records: list[EntityRecord], output_dir: Path, job_id: str):
    audit_path = output_dir / f"{job_id}_audit.jsonl"
    with open(audit_path, "w") as f:
        for rec in records:
            f.write(json.dumps(rec.model_dump(mode="json")) + "\n")
    logger.info(f"Audit trail written: {audit_path}")
    return audit_path
